from flask import Flask, request, jsonify
from flask_cors import CORS
import pandas as pd
import numpy as np
import traceback
import matplotlib.pyplot as plt
import seaborn as sns
import base64
from io import BytesIO
from scipy.signal import butter, filtfilt
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix

# ✅ REQUIRED IMPORTS FOR EXCEL
import os
from openpyxl import Workbook, load_workbook

plt.switch_backend("Agg")

app = Flask(__name__)
CORS(app)

# =====================================================
# Convert plot to Base64
# =====================================================
def fig_to_base64(fig):
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    buf.seek(0)
    img = base64.b64encode(buf.getvalue()).decode("utf-8")
    plt.close(fig)
    return "data:image/png;base64," + img


# =====================================================
# Band-pass filter
# =====================================================
def bandpass_filter(data, lowcut, highcut, fs=500, order=4):
    nyq = 0.5 * fs
    b, a = butter(order, [lowcut / nyq, highcut / nyq], btype="band")
    return filtfilt(b, a, data)


# =====================================================
# EEG Prediction API
# =====================================================
@app.route("/api/predict", methods=["POST"])
def predict():
    try:
        file = request.files.get("file")
        if not file:
            return jsonify({"error": "No file uploaded"}), 400

        df = pd.read_csv(file)
        if df.empty:
            return jsonify({"error": "Empty CSV"}), 400

        label_col = "Cognitive_State" if "Cognitive_State" in df.columns else df.columns[-1]

        eeg_channels = [
            c for c in df.columns
            if c != label_col and pd.api.types.is_numeric_dtype(df[c])
        ]

        if not eeg_channels:
            return jsonify({"error": "No EEG channel columns found"}), 400

        filtered = df[eeg_channels].astype(float)

        fs = 500
        window_size = fs
        n_windows = len(filtered) // window_size

        features, labels = [], []

        for w in range(n_windows):
            s, e = w * window_size, (w + 1) * window_size
            win = filtered.iloc[s:e]

            if len(win) == 0:
                continue

            labels.append(df[label_col].iloc[s:e].mode().iloc[0])
            features.append(win.mean().to_dict())

        if not features:
            return jsonify({"error": "Not enough data"}), 400

        features_df = pd.DataFrame(features)
        features_df["label"] = labels

        X = features_df.drop(columns=["label"]).values
        y = features_df["label"].values

        le = LabelEncoder()
        y_enc = le.fit_transform(y)

        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)

        X_train, X_test, y_train, y_test = train_test_split(
            X_scaled, y_enc, test_size=0.2, random_state=42
        )

        clf = RandomForestClassifier(n_estimators=200, random_state=42)
        clf.fit(X_train, y_train)
        y_pred = clf.predict(X_test)

        acc = accuracy_score(y_test, y_pred)
        report = classification_report(y_test, y_pred, target_names=le.classes_)
        cm = confusion_matrix(y_test, y_pred)

        fig = plt.figure(figsize=(5, 4))
        sns.heatmap(cm, annot=True, fmt="d", cmap="Blues",
                    xticklabels=le.classes_, yticklabels=le.classes_)
        plt.title("Confusion Matrix")
        cm_img = fig_to_base64(fig)

        return jsonify({
            "accuracy": acc,
            "classification_report": report,
            "confusion_matrix": cm.tolist(),
            "images": {
                "confusion_matrix": cm_img
            }
        })

    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


# =====================================================
# CONTACT FORM → SAVE TO EXCEL
# =====================================================
EXCEL_FILE = "contact_data.xlsx"

def save_contact_to_excel(name, email, message):
    file_path = os.path.abspath(EXCEL_FILE)
    print("📁 Saving to:", file_path)

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contact Data"
        ws.append(["Name", "Email", "Message"])
        wb.save(file_path)

    wb = load_workbook(file_path)
    ws = wb.active
    ws.append([name, email, message])
    wb.save(file_path)

    print("✅ Contact data saved successfully")


@app.route("/api/contact", methods=["POST"])
def contact():
    try:
        print("✅ /api/contact API HIT")

        data = request.get_json()
        print("📦 Data received:", data)

        name = data.get("name")
        email = data.get("email")
        message = data.get("message")

        if not name or not email or not message:
            return jsonify({"message": "All fields are required"}), 400

        save_contact_to_excel(name, email, message)

        return jsonify({"message": "Contact details saved successfully!"})

    except PermissionError:
        return jsonify({
            "error": "Excel file is open. Close contact_data.xlsx and try again."
        }), 500

    except Exception as e:
        print("❌ CONTACT ERROR")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(port=5000, debug=True)
